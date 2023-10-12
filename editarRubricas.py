import tkinter as tk
import moduloScroll, openpyxl
from openpyxl import load_workbook
from tkinter import filedialog
from tkinter import ttk



def main(root, scrollable_frame):
    frameRubrica = ttk.Frame(scrollable_frame)
    frameRubrica.pack()
    global entry_count_column, entry_count_row

    # Abre el la rúbirca
    rutaRubrica = filedialog.askopenfilename(title="Elige una rúbrica para modificar", initialdir="./_internal/Rubricas/",
                                             filetypes=[("Archivos xlsx", "*.xlsx")])

    # Crear los frames
    frameButton = ttk.Frame(frameRubrica)
    frameButton.grid(row=0, padx=10, pady=10)
    frameLabel = ttk.Frame(frameRubrica)
    frameLabel.grid(row=1, padx=10, pady=10)
    frame = ttk.Frame(frameRubrica)
    frame.grid()
    row_frame = ttk.Frame(frame)
    row_frame.grid()
    
    # Carga la rúbrica a modificar
    rubrica = load_workbook(rutaRubrica)
    nombreHoja = rubrica.sheetnames[0]
    hoja = rubrica[nombreHoja]
    rubrica.active

    inforNombre = ttk.Label(frameButton, text="Rúbrica a moficar:")
    inforNombre.grid(row=2, column=0)
    nombreRubrica = ttk.Label(frameButton, text=f"{rutaRubrica}")
    nombreRubrica.grid(row=2,column=1, padx=10, pady=5)
    nombreRubrica.configure(width=60)
      
    # Variables
    entry_count_column = hoja.max_column
    entry_count_row = hoja.max_row
    print(entry_count_column, entry_count_row)
    
    def add_row(): #ARREGLAR ESTO
        global entry_count_row
        entry_count_row +=1
        # Crear cuadros de texto dispuestos de forma horizontal en filas
        row_frame = ttk.Frame(frame)
        row_frame.grid(row=entry_count_row)

        for j in range(entry_count_column):  # Cambiar el número de cuadros de texto por fila según sea necesario
            text = tk.Text(row_frame, height=10, width=30)
            text.grid(row=0, column=j, padx=10, pady=5)

     # Estilo de los widgets
    style = ttk.Style()
    style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
              background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
    
    ttk.Style().configure("TFrame", background="#649BEF")
    ttk.Style().configure("TLabel", background="#649BEF")
    ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                           borderwidth=0, font=(18), weigth="blod", bordercolor="blue", weight='bold')


    def save_to_excel(frame, file_path):
        hojaNueva = []
        wb = openpyxl.Workbook()
        hojaNueva = wb.active
        
        # selecciona el texto de los Text
        for row in frame.winfo_children():    
            hojaNueva.append([entry.get("1.0", "end-1c") for entry in row.winfo_children() if isinstance(entry,
             tk.Text)])
        # Guardar el array en un archivo Excel
        wb.save(rutaRubrica)



        save = ttk.Label(frameLabel, text=f"La rúbrica se ha guardado correctamente en {rutaRubrica}",
                        fg="green")
        save.grid(row=3, padx=10, pady=5)
    


#--------------Botones de la ventana ---------------------
    # Crear fila de rúbricas
    add_fila = ttk.Button(frameButton, text="Añadir rúbrica", command=add_row)
    add_fila.grid(row=0, column=1, padx=10, pady=5)

    # Crear el botón para guardar en un archivo CSV
    save_button = ttk.Button(frameButton, text="Guardar", command=lambda: save_to_excel(frame, rutaRubrica))
    save_button.grid(row=0, column=2, padx=10, pady=5)
 
    for j in range(entry_count_column):  # Cambiar el número de cuadros de texto por fila según sea necesario
            text = tk.Text(row_frame, height=2, width=30)
            text.grid(row=1, column=j, padx=10, pady=5)
            celda = hoja.cell(row=1, column=j+1)
            valorCelda = celda.value
            text.insert(tk.END, valorCelda)
    
    
    criterio = ttk.Label(frameLabel, text="En la primera casilla debes escribir 'Criterios de Evaluación'")
    criterio.grid(row=0, padx=10, pady=5)
    info = ttk.Label(frameLabel, text="En las siguientes casillas de la primera fila debes escribir "\
                    "cómo calificas la rúbrica. Excelente, bien, etc.")
    info.grid(row=1, padx=10, pady=5)


    # Crear los cinco Entry iniciales
    for fila in range(2,entry_count_row+1):
        #global entry_count_row
        entry_count_row +=1
        # Crear cuadros de texto dispuestos de forma horizontal en filas
        row_frame = ttk.Frame(frame)
        row_frame.grid()

        for j in range(entry_count_column):  # Cambiar el número de cuadros de texto por fila según sea necesario
            text = tk.Text(row_frame, height=10, width=30)
            text.grid(row=fila, column=j, padx=10, pady=5)
            celda = hoja.cell(row=fila, column=j+1)
            valorCelda = celda.value
            text.insert(tk.END, valorCelda)
    
    

    root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Grupo de alumnos")
    root.geometry("700x1000")
    
    #---------- FRAME PARA METER EL CONTENIDO -------------------------------
    scrollable_frame = moduloScroll.ScrollableFrame(root)
    scrollable_frame.pack(fill="both", expand=True)

    main(root, scrollable_frame.scrollable_frame)

    root.mainloop()
