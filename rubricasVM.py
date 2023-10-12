import tkinter as tk
from tkinter import ttk
from tkinter import *
from tkinter.messagebox import *
import moduloScroll, subprocess, openpyxl, creaTabs, os, creaTabs, editarGrupos, editarRubricas, editarSituacion
import grupoAlumnos,informes, rubricaNueva, situacion, webbrowser
from openpyxl.styles import Font
from tkinter import filedialog
        
def main():
    root = tk.Tk()
    root.title("Rúbricas VM")
    root.geometry("1700x900")
    root.configure(bg="#649BEF")
    #---------- FRAME PARA METER EL CONTENIDO -------------------------------
    scrollable_frame = moduloScroll.ScrollableFrame(root)
    scrollable_frame.pack(fill="both", expand=True)
    

     #-------------- MENÚ DE LA APLICACIÓN ------------------
    def nuevaRubrica():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        
        rubricaNueva.main(root, scrollable_frame.scrollable_frame)

    def nuevoGrupoAlumnos():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        
        grupoAlumnos.main(root, scrollable_frame.scrollable_frame)

    def nuevaSituacion():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        
        situacion.main(root, scrollable_frame.scrollable_frame)
       
    def abrirSituacion():
        rutaSituacion = filedialog.askopenfilename(title="Elige una Situación de aprendizaje",
                                                   initialdir="./_internal/SituacionesAprendizaje/",
                                                   filetypes=[("Archivos xlsx", "*.xlsx",)])

        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        # Lógica para abrir un archivo
        situacionAprendizaje = openpyxl.load_workbook(rutaSituacion)
        botones(situacionAprendizaje, rutaSituacion)

    def editarRubrica():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        
        editarRubricas.main(root, scrollable_frame.scrollable_frame)
        
    def editarGrupo():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        
        editarGrupos.main(root, scrollable_frame.scrollable_frame)

    def editarSituaciones():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        
        editarSituacion.main(root, scrollable_frame.scrollable_frame)
        

    def informeAlumno():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()
        
        informes.main(root, scrollable_frame.scrollable_frame)

    def ayuda():
        ventana = scrollable_frame.scrollable_frame.winfo_children()
        ventanaCerar = ventana[0]
        ventanaCerar.destroy()

        frameAyuda = ttk.Frame(scrollable_frame.scrollable_frame)
        frameAyuda.pack()

        # Estilo de los widgets
        style = ttk.Style()
        style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
                  background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
        ttk.Style().configure("TFrame", background="#649BEF")
        ttk.Style().configure("TLabel", background="#649BEF", font=(18))
        ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                               borderwidth=0, font=(18), weigth="blod", bordercolor="blue", weight='bold')
        
        fuente = ("Arial", 20, "bold")
        fuenteb = ("Arial", 15, "bold")
        fuentec = ("Arial", 12)

        nombreApp = ttk.Label(frameAyuda, text="VMRúbricas", font=fuente)
        nombreApp.grid(row=0,column=0, padx=40, pady=5)
        nombreApp.configure(foreground="purple")

        texto = " VMRúbricas es una aplicación para evaluar alumnos mediante rúbricas"
        textoApp = ttk.Label(frameAyuda, text=texto, font=fuenteb)
        textoApp.grid(row=1,column=0, padx=40, pady=5)

        textoA = "   Con ella puedes crear rúbricas o utilizar otras ya hechas, asignarlas a un grupo y evaluar facilmente en la clase, además se puede ver la evaluación de los alumnos de forma individual\n o colectiva y guardar los informes de evaluación en formato excell o pdf."
        textoAApp = ttk.Label(frameAyuda, text=textoA, font=fuentec)
        textoAApp.grid(row=2,column=0, padx=40, pady=5)

        textoA = "   Para más información y ver el funcionamiento puedes visitar el siguiente enlace:"
        textoAApp = ttk.Label(frameAyuda, text=textoA, font=fuentec)
        textoAApp.grid(row=3,column=0, padx=40, pady=5, sticky='w')

        enlace = ttk.Button(frameAyuda, text="Web de VMRúbricas", command=lambda: webbrowser.open_new("https://www.educa2.madrid.org/web/victormanuel.patonmancebo/vm-rubricas"))
        enlace.grid(row=4, column=0, padx=20, pady=30)

        licencia = ttk.Button(frameAyuda, text="Licencia de uso Creative Commons: CC BY-NC-SA 4.0 Internacional",
                               command=lambda: webbrowser.open_new("https://creativecommons.org/licenses/by-nc-sa/4.0/deed.es"))
        licencia.grid(row=5, column=0, padx=40, pady=30, sticky='w')




    def acerca():
        showinfo(title="VMRúbricas", message="Rúbricas VM. Versión: 1.0")

    # Función para salir del programa
    def salir():
        root.quit()

    # Barra de menú
    barra_menu = tk.Menu(root)
    root.config(menu=barra_menu)

    # Menú "Archivo"
    menu_archivo = tk.Menu(barra_menu, tearoff=0)
    barra_menu.add_cascade(label="Archivo", menu=menu_archivo)
    menu_archivo.add_command(label="Nueva rúbrica", command=nuevaRubrica)
    menu_archivo.add_command(label="Nuevo grupo de alumnos", command=nuevoGrupoAlumnos)
    menu_archivo.add_command(label="Nueva situación de aprendizaje", command=nuevaSituacion)
    menu_archivo.add_command(label="Abrir situación de aprendizaje", command=abrirSituacion)
    menu_archivo.add_separator()
    menu_archivo.add_command(label="Salir", command=salir)

    # Menú "Editar"
    menu_editar = tk.Menu(barra_menu, tearoff=0)
    barra_menu.add_cascade(label="Editar", menu=menu_editar)
    menu_editar.add_command(label="Editar rúbrica", command=editarRubrica)
    menu_editar.add_command(label="Editar grupo de alumnos", command=editarGrupo)
    menu_editar.add_command(label="Editar alumnos en situación de aprendizaje", command=editarSituaciones)

    # Menú "Herramientas"
    menu_herramientas = tk.Menu(barra_menu, tearoff=0)
    barra_menu.add_cascade(label="Informes", menu=menu_herramientas)
    menu_herramientas.add_command(label="Informes de los alumnos", command=informeAlumno)

    # Menú "Ayuda"
    menu_ayuda = tk.Menu(barra_menu, tearoff=0)
    barra_menu.add_cascade(label="Ayuda", menu=menu_ayuda)
    menu_ayuda.add_command(label="Ayuda", command=ayuda)
    menu_ayuda.add_command(label="Acerca de VMRúbricas", command=acerca)

    #---------- FIN DE MENÚ ----------------------------

    
    #--------- frame de información --------------------------------------------
    infoFrame = tk.Frame(scrollable_frame.scrollable_frame)


    infoFrame.pack()
    infoFrame.configure(background="#649BEF")
    infoLabel = tk.Label(infoFrame, text="INSTRUCCIONES: - Para configurar la evaluación de una situación de aprendizaje debes configurar primero al menos un grupo y una rúbrica.",justify='center')
    infoLabel.configure(background="#649BEF")
    infoLabel.grid(row=0, column=0, padx=20,pady=20, rowspan=1)
    fuente = ("Helvetica", 500, "bold")
    fuenteB = ("Arial", 100, "bold")
    nombreApp = ttk.Label(infoFrame, text="VM", font=fuente)
    nombreApp.grid(row=2,column=0, padx=40, pady=0)
    nombreApp.configure(foreground="purple")
    nombreApplittle = ttk.Label(infoFrame, text="Rúbricas", font=fuenteB )
    nombreApplittle.grid(row=1, column=0, padx=40, pady=0)
    nombreApplittle.configure(justify='center')
    
    
    style = ttk.Style()
    style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
              background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
    
    ttk.Style().configure("TFrame", background="#649BEF")
    ttk.Style().configure("TLabel", background="#649BEF", foreground="white")
    ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                           borderwidth=1, size= 18, weigth="blod")
    

    def botones(situacionAprendizaje, rutaSituacion):
        notebook = ttk.Notebook(scrollable_frame.scrollable_frame)
        notebook.grid(row=0, column=0)
        infoFrame.destroy()
        navigation_frame = ttk.Frame(notebook)
        notebook.add(navigation_frame, text="ALUMNOS")
        nombreArchivo = os.path.basename(rutaSituacion)
        informacion = ttk.Label(navigation_frame, text=f"Situación de aprendizaje: {nombreArchivo} " )
        informacion.grid(row=0, column=2, padx=20, pady=5)
        informacion.configure(foreground='black')

        archivo = openpyxl.load_workbook(rutaSituacion)

        # Selecciona la hoja que desees procesar (por ejemplo, la primera hoja)
        hoja = archivo.active
        nombresHojas = archivo.sheetnames
        nombreHoja = nombresHojas[1]
        hojaCriterios = archivo[nombreHoja]

        contador = 1       # Recorre la primera columna e imprime el contenido de cada celda

        for fila in hojaCriterios.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
            for celda in fila:
                etiquetaCriterio = ttk.Label(navigation_frame, text=f"{celda}")
                etiquetaCriterio.grid(row=contador, column=2)
                etiquetaCriterio.configure(foreground="black")
                contador +=1

        # Frame dentro del Canvas para los botones
        nombresAlumnos = situacionAprendizaje.sheetnames
        buttons_frame = ttk.Frame(navigation_frame)
        buttons_frame.configure()
        numeroAlumnos = len(nombresAlumnos)
        for i in range(1,numeroAlumnos):
            nombre = nombresAlumnos[i]
            ttk.Button(navigation_frame,
                        text=f"{nombre}", command=lambda index=i: notebook.select(index)).grid(row=i,column=0)
            
        creaTabs.create_tabs(notebook, situacionAprendizaje, nombresAlumnos, rutaSituacion)

    root.mainloop()
    


main()
    