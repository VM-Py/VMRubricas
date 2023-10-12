import tkinter as tk
from tkinter import ttk
from tkinter import *
import openpyxl, moduloScroll, os
from tkinter import filedialog
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


def main(root, scrollable_frame):
    #---------- FRAME PARA METER EL CONTENIDO -------------------------------
    frameInforme = ttk.Frame(scrollable_frame)
    frameInforme.pack()
    
    informesFrame = ttk.Frame(frameInforme)
    informesFrame.grid(padx=10, pady=5, row=1, column=0, columnspan=1)

    style = ttk.Style()
    style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
              background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
    ttk.Style().configure("TFrame", background="#649BEF")
    ttk.Style().configure("TLabel", background="#649BEF", font=(18))
    ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                           borderwidth=0, font=(18), weigth="blod", bordercolor="blue", weight='bold')

    rutaSituacion = filedialog.askopenfilename(title="Elige una situación para ver los informes",
                                                initialdir="./_internal/SituacionesAprendizaje/",
                                           filetypes=[("Archivos xlsx", "*.xlsx")])
    nombreArchivo = os.path.basename(rutaSituacion)
    
    situacion = openpyxl.load_workbook(rutaSituacion)
    # Selecciona la hoja que desees procesar (por ejemplo, la primera hoja)
    hoja = situacion.active

    def informesDeRecuperacion():
        nombresAlumnos = situacion.sheetnames
        wb = openpyxl.Workbook()
        for alumno in nombresAlumnos[1:]:
            letraColumna = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
            informe = []
            hoja = situacion[f"{alumno}"]
            numeroColumnas = hoja.max_column
            numeroFilas = hoja.max_row
            informe = wb.active
            informe.append(["Rúbrica elaborada con VM Rúbricas","."])
            informe.append(["Situación: " ,nombreArchivo])
            informe.append(["Alumno",alumno])

            notas = 0
            nota = 0
            notab = 0
            notasb = 0
            for fila in range(1, numeroFilas+1):
                for columna in range(0, numeroColumnas):
                    col = letraColumna[columna]
                    celda = hoja[col+str(fila)]
                    valorCelda = celda.value

                    if columna == 0:
                        celdaCriterio = hoja[col+str(fila)]
                        valorCeldaCriterio = celdaCriterio.value

                    fuenteCelda = celda.font
                    color_texto = fuenteCelda.color
                    celdaVerde = 'FF08CB43'        
                    if color_texto is not None and color_texto.rgb == celdaVerde or color_texto.rgb == "0008CB43":
                        if columna == numeroColumnas-1:
                            informe.append([valorCeldaCriterio, valorCelda])
                        nota = numeroColumnas-columna
                        for contador in range(1, numeroColumnas):
                            if columna == numeroColumnas-1:
                                notab = 1
                            elif columna >1 and columna<=numeroColumnas-2:
                                notab = 5
                            elif columna >numeroColumnas-2 and columna<=numeroColumnas-3:
                                notab = 7
                            else:
                                notab = 10

                notas = notas+nota
                notasb = notasb+notab
            calificacion = notas/(numeroColumnas-1)
            calificacionb = notasb/(numeroColumnas-1)
            informe.append(["Calificacion sobre el número de items:", calificacion])
            informe.append(["Calificacion sobre 10:", calificacionb])
            informe.append(["-----------------------------------------------","-----------------"])

        columna = informe['A']
        valores = [celda.value for celda in columna]
        fila = 1
        for valor in valores:
            if valor != '':
               informe.cell(row=fila, column=1, value=valor)
               fila +=1
        nombreArchivoGuardar = f"Informe de recuperación del grupo {nombreArchivo}"
        wb.save(f"./_internal/Informes/{nombreArchivoGuardar}")
        infoGuardar = ttk.Label(informesFrame, text=f"Se ha guardado en ./_internal/Informes/{nombreArchivoGuardar}")
        infoGuardar.grid(row=3, column=4, padx=30,pady=5)
        infoGuardar.configure(foreground="green")    
        

    def informesAllAlumnos():
        nombresAlumnos = situacion.sheetnames
        wb = openpyxl.Workbook()
        calificacionesb =[]
        alumnoConPendiente = 0
        numFilas = 0

        for alumno in nombresAlumnos[1:]:
            letraColumna = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

            pendiente = False
            alumnoSuspenso = 0
            alumnoSuficiente = 0
            alumnoBien = 0
            alumnoNotable = 0
            alumnoSobresaliente = 0
            sinCalificar = 0

            informe = []
            hoja = situacion[f"{alumno}"]
            numeroColumnas = hoja.max_column
            numeroFilas = hoja.max_row
            informe = wb.active
            informe.title = "Informe de grupo"
            informe.append(["Rúbrica elaborada con VM Rúbicas","."])
            informe.append(["Situación: " ,nombreArchivo])
            informe.append(["Alumno",alumno])

            notas = 0
            nota = 0
            notab = 0
            notasb = 0

            for fila in range(1, numeroFilas+1):
                for columna in range(0, numeroColumnas):
                    col = letraColumna[columna]
                    celda = hoja[col+str(fila)]
                    valorCelda = celda.value

                    if columna == 0:
                        celdaCriterio = hoja[col+str(fila)]
                        valorCeldaCriterio = celdaCriterio.value

                    fuenteCelda = celda.font
                    color_texto = fuenteCelda.color
                    celdaVerde = 'FF08CB43'        
                    if color_texto is not None and color_texto.rgb == celdaVerde or color_texto.rgb == "0008CB43":
                        informe.append([valorCeldaCriterio, valorCelda])
                        print(valorCeldaCriterio, valorCelda)
                        nota = numeroColumnas-columna
                        for contador in range(1, numeroColumnas):
                            if columna == numeroColumnas-1:
                                notab = 1
                                pendiente=True
                                
                                redCell = informe['B'+str((fila+2)+numFilas)]
                                font = Font(color="FF0000")
                                redCell.font = font

                            elif columna >1 and columna<=numeroColumnas-2:
                                notab = 5
                            elif columna >numeroColumnas-2 and columna<=numeroColumnas-3:
                                notab = 7
                            else:
                                notab = 10

                notas = notas+nota
                notasb = notasb+notab

            calificacion = notas/(numeroColumnas-1)
            calificacionb = notasb/(numeroColumnas-1)
            calificacionesb.append(calificacionb)
            
            informe.append(["Calificacion sobre el número de items:", calificacion])
            informe.append(["Calificacion sobre 10:", calificacionb])
            informe.append(["-----------------------------------------------","-----------------"])
            print("numFilas: ", numFilas)
            if calificacion == 0:
                numFilas = numFilas+6
            else:
                numFilas+=5+numeroFilas

            
            if pendiente == True:
                alumnoConPendiente +=1
        
        columna = informe['A']
        valores = [celda.value for celda in columna]
        fila = 1
        for valor in valores:
            if valor != '':
               informe.cell(row=fila, column=1, value=valor)
               fila +=1
        nombreArchivoGuardar = f"Informe de todos los alumnos de {nombreArchivo}"

        for califica in calificacionesb:
            if califica == 0:
                sinCalificar +=1
            elif califica<5:
                alumnoSuspenso+=1
            elif califica<=6:
                alumnoSuficiente +=1
            elif calificacionb<7:
                alumnoBien+=1
            elif califica<9:
                alumnoNotable+=1
            else:
                alumnoSobresaliente+=1

        numeroAlumnos = len(nombresAlumnos)
        aprobados = alumnoSuficiente+alumnoBien+alumnoNotable+alumnoSobresaliente
        porcentajeAprobados = aprobados*100/numeroAlumnos

        estadisticas = wb.create_sheet("Estadísticas")
        estadisticas.cell(row=1, column=1, value="ESTADÍSTICAS")
        estadisticas.cell(row=1, column=2, value=f"SinCalificar:{sinCalificar}")
        estadisticas.cell(row=2, column=1, value="Número de suspensos:")
        estadisticas.cell(row=2, column=2, value=alumnoSuspenso)
        estadisticas.cell(row=3, column=1, value="Número de alumnos con Suficiente:")
        estadisticas.cell(row=3, column=2, value=alumnoSuficiente)
        estadisticas.cell(row=4, column=1, value="Número de alumnos con Bien :")
        estadisticas.cell(row=4, column=2, value=alumnoBien)
        estadisticas.cell(row=5, column=1, value="Número de alumnos con Notable:")
        estadisticas.cell(row=5, column=2, value=alumnoNotable)
        estadisticas.cell(row=6, column=1, value="Número de alumnos con Suficiente:")
        estadisticas.cell(row=6, column=2, value=alumnoSobresaliente)
        estadisticas.cell(row=8, column=1, value="Porcentaje de aprobados")
        estadisticas.cell(row=8, column=1, value="Porcentaje de aprobados")
        estadisticas.cell(row=8, column=2, value=f"{porcentajeAprobados}%")
        estadisticas.cell(row=10, column=1, value="Número de Alumnos")
        estadisticas.cell(row=10, column=2, value=f"{numeroAlumnos-1}")
        estadisticas.cell(row=11, column=1, value="Alumnos con criterios pendientes")
        estadisticas.cell(row=11, column=2, value=f"{alumnoConPendiente}")
        
        

        wb.save(f"./_internal/Informes/{nombreArchivoGuardar}")
        infoGuardar = ttk.Label(informesFrame, text=f"Se ha guardado en ./_internal/Informes/{nombreArchivoGuardar}")
        infoGuardar.grid(row=3, column=4, padx=30,pady=5)
        infoGuardar.configure(foreground="green")    
            

    def recuperacionAlumno(button):
        antiguoFrame = frameInforme.winfo_children()
        if len(antiguoFrame)>2:
            borrarFrame = antiguoFrame[2]
            borrarFrame.destroy()
        
       
        informeFrame = tk.Frame(frameInforme)
        informeFrame.grid(row=2, column=4)
        informeFrame.configure(background='#C7E5FC')
         # Estilo de los widgets
        style = ttk.Style()
        style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
                  background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
        ttk.Style().configure("TFrame", background="#649BEF")
        ttk.Style().configure("TLabel", background="#649BEF", font=(18))
        ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                               borderwidth=0, font=(18), weigth="blod", bordercolor="blue", weight='bold')
        alumno = button[0]

        hoja = situacion[f"{alumno}"]
        numeroColumnas = hoja.max_column
        numeroFilas = hoja.max_row
        letraColumna = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        
        nombreSituacion = ttk.Label(informeFrame, text=f"Situación de aprendizaje: {nombreArchivo}")
        nombreSituacion.grid(row=0,column=0, padx=10, pady=5, sticky="w")
        nombreSituacion.configure(background='#C7E5FC', foreground='black')
        nombreLabel = ttk.Label(informeFrame, text=f"Informe del alumno: {alumno}")
        nombreLabel.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        nombreLabel.configure(background='#C7E5FC', foreground='black')
        
        informe = []
        wb = openpyxl.Workbook()
        informe = wb.active
        informe.append(["Rúbrica elaborada con VM Rúbicas","."])
        informe.append(["Situación: " ,nombreArchivo])
        informe.append(["Alumno",alumno])


        for fila in range(1, numeroFilas+1):
            for columna in range(0, numeroColumnas):
                col = letraColumna[columna]
                celda = hoja[col+str(fila)]
                valorCelda = celda.value
                
                if columna == 0:
                    celdaCriterio = hoja[col+str(fila)]
                    valorCeldaCriterio = celdaCriterio.value

                fuenteCelda = celda.font
                color_texto = fuenteCelda.color      
                if columna == numeroColumnas-1 and color_texto is not None and color_texto.rgb == "0008CB43":
                    criterio = ttk.Label(informeFrame, text=f"{valorCeldaCriterio}")
                    criterio.grid(row=fila+1, column=0, padx=10, pady=5)
                    criterio.configure(background='#C7E5FC', foreground='black')
                    informe.append([valorCeldaCriterio, valorCelda])
                    itemAlumno = ttk.Label(informeFrame, text=valorCelda)
                    itemAlumno.grid(row=fila+1, column=1, padx=10, pady=5, sticky="w")
                    itemAlumno.configure(foreground="red", background='#C7E5FC')
        
        def exportarAExcel():
            columna = informe['A']
            valores = [celda.value for celda in columna]
            fila = 1
            for valor in valores:
                if valor != '':
                   informe.cell(row=fila, column=1, value=valor)
                   fila +=1

            nombreArchivoGuardar = f"Informe de recuperacion del {alumno} en {nombreArchivo}"
            wb.save(f"./_internal/Informes/{nombreArchivoGuardar}")
            infoGuardar = ttk.Label(informeFrame, text=f"Se ha guardado en ./_internal/Informes/{nombreArchivoGuardar}")
            infoGuardar.grid(row=5, column=3, padx=30,pady=5)
            infoGuardar.configure(foreground="green", background='#C7E5FC')

        def exportarPdf():
            columna = informe['A']
            valores = [celda.value for celda in columna]
            fila = 1
            for valor in valores:
                if valor != '':
                   informe.cell(row=fila, column=1, value=valor)
                   fila +=1

            nombreArchivoGuardar = f"Informe de recuperación de {alumno} en {nombreArchivo}"
            infoGuardar = ttk.Label(informeFrame, text=f"Se ha guardado {nombreArchivoGuardar}")
            infoGuardar.grid(row=5, column=3, padx=30,pady=5)
            infoGuardar.configure(foreground="green", background='#C7E5FC')

            hoja = wb.active  # Puedes cambiar esto para seleccionar una hoja específica

            # Configura el lienzo PDF
            c = canvas.Canvas(f"./_internal/Informes/{nombreArchivoGuardar}.pdf", pagesize=letter)

            # Configura el tamaño de la fuente y el espacio entre líneas
            #c.setFont("Arial", 12)
            espacio_entre_lineas = 14

            # Inicializa las coordenadas para dibujar en la página
            x = 100
            y = 750

            # Itera a través de las filas y columnas del archivo XLSX
            for fila in hoja.iter_rows():
                x = 100
                for celda in fila:
                    # Obtiene el valor de la celda
                    valor = celda.value
                        # # Dibuja el valor en el PDF
                    c.drawString(x, y, str(valor))  # Dibuja en la posición actual (x, y)
                    x+=200
                # Mueve la posición vertical hacia arriba
                y -= espacio_entre_lineas

            # Guarda el PDF
            c.save()

        exportarButton = ttk.Button(informeFrame, text="Guardar en anchivo", command=exportarAExcel)
        exportarButton.grid(row=0, column=3, padx=30, pady=5)
        exportarPdf = ttk.Button(informeFrame, text="Guardar en PDF", command=exportarPdf)
        exportarPdf.grid(row=1, column=3, padx=30, pady=5)


    def informeAlumno(button):
        antiguoFrame = frameInforme.winfo_children()
        if len(antiguoFrame)>2:
            borrarFrame = antiguoFrame[2]
            borrarFrame.destroy()
            
        informeFrame = tk.Frame(frameInforme)
        informeFrame.grid(row=2, column=4)
        informeFrame.configure(background='#C7E5FC')
         # Estilo de los widgets
        style = ttk.Style()
        style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
                  background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
        ttk.Style().configure("TFrame", background="#649BEF")
        ttk.Style().configure("TLabel", background="#649BEF", foreground="white")
        ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                           borderwidth=0, size= 18, weigth="blod")
        fuente_negrita = ("Helvetica", 12, "bold")
        alumno = button[0]

        hoja = situacion[f"{alumno}"]
        numeroColumnas = hoja.max_column
        numeroFilas = hoja.max_row
        letraColumna = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        
        nombreSituacion = ttk.Label(informeFrame, text=f"Situación de aprendizaje: {nombreArchivo}", font=fuente_negrita)
        nombreSituacion.grid(row=0,column=0, padx=10, pady=5, sticky="w")
        nombreSituacion.configure(background='#C7E5FC', foreground='black')
        nombreLabel = ttk.Label(informeFrame, text=f"Informe del alumno: {alumno}", font=fuente_negrita)
        nombreLabel.grid(row=1, column=0, padx=10, pady=5, sticky="w")
        nombreLabel.configure(background='#C7E5FC', foreground='black')
        
        informe = []
        wb = openpyxl.Workbook()
        informe = wb.active
        informe.append(["Rúbrica elaborada con VM Rúbricas","."])
        informe.append(["Situación: " ,nombreArchivo])
        informe.append(["Alumno",alumno])

        notas = 0
        nota = 0
        notab = 0
        notasb = 0
        for fila in range(1, numeroFilas+1):
            for columna in range(0, numeroColumnas):
                col = letraColumna[columna]
                celda = hoja[col+str(fila)]
                valorCelda = celda.value
                
                if columna == 0:
                    celdaCriterio = hoja[col+str(fila)]
                    valorCeldaCriterio = celdaCriterio.value
                    criterio = ttk.Label(informeFrame, text=f"{valorCeldaCriterio}")
                    criterio.grid(row=fila+1, column=0, padx=10, pady=5)
                    criterio.configure(background='#C7E5FC', foreground='black')

                fuenteCelda = celda.font
                color_texto = fuenteCelda.color
                celdaVerde = 'FF08CB43'      
                if color_texto is not None and color_texto.rgb == celdaVerde or color_texto.rgb == "0008CB43":
                    informe.append([valorCeldaCriterio, valorCelda])

                    if columna == numeroColumnas-1:
                        itemAlumno = ttk.Label(informeFrame, text=valorCelda)
                        itemAlumno.grid(row=fila+1, column=1, padx=10, pady=5, sticky="w")
                        itemAlumno.configure(foreground="red", background='#C7E5FC')
                        redCell = informe['B'+str(fila+2)]
                        font = Font(color="FF0000")
                        redCell.font = font

                    else:    
                        itemAlumno = ttk.Label(informeFrame, text=valorCelda)
                        itemAlumno.grid(row=fila+1, column=1, padx=10, pady=5, sticky="w")
                        itemAlumno.configure(foreground="green", background='#C7E5FC')
                        greenCell = informe['B'+str(fila+2)]
                        font = Font(color="26C607")
                        greenCell.font = font
                    nota = numeroColumnas-columna
                    for contador in range(1, numeroColumnas):
                        if columna == numeroColumnas-1:
                            notab = 1
                        elif columna >1 and columna<=numeroColumnas-2:
                            notab = 5
                        elif columna >numeroColumnas-2 and columna<=numeroColumnas-3:
                            notab = 7
                        else:
                            notab = 10
                        
                    print(columna, notab)
            notas = notas+nota
            notasb = notasb+notab    
        calificacion = notas/(numeroColumnas-1)
        calificacionb = notasb/(numeroColumnas-1)
        informe.append(["Calificacion:", calificacion])
        muestraNota = ttk.Label(informeFrame, text=f"Calificación: {calificacion} sobre {numeroColumnas-1}", 
                  font=fuente_negrita)
        muestraNota.grid(column=0, padx=10, pady=5)
        muestraNota.configure(background='#C7E5FC', foreground='black')
        informe.append(["Calificacion:", calificacionb])
        muestraNotaDiez =ttk.Label(informeFrame, text=f"Calificación: {calificacionb} sobre 10", 
                  font=fuente_negrita)
        muestraNotaDiez.grid(column=0, padx=10, pady=5)
        muestraNotaDiez.configure(background='#C7E5FC', foreground='black')
        
        
        

        def exportarAExcel():
            columna = informe['A']
            valores = [celda.value for celda in columna]
            fila = 1
            for valor in valores:
                if valor != '':
                   informe.cell(row=fila, column=1, value=valor)
                   fila +=1

            nombreArchivoGuardar = f"{alumno}_{nombreArchivo}"
            wb.save(f"./_internal/Informes/{nombreArchivoGuardar}")
            infoGuardar = ttk.Label(informeFrame, text=f"Se ha guardado en ./_internal/Informes/{nombreArchivoGuardar}")
            infoGuardar.grid(row=5, column=3, padx=30,pady=5)
            infoGuardar.configure(foreground="green", background='#C7E5FC')

        def exportarPdf():
            columna = informe['A']
            valores = [celda.value for celda in columna]
            fila = 1
            for valor in valores:
                if valor != '':
                   informe.cell(row=fila, column=1, value=valor)
                   fila +=1
                   

            nombreArchivoGuardar = f"{alumno}_{nombreArchivo}"
            infoGuardar = ttk.Label(informeFrame, text=f"Se ha guardado {nombreArchivoGuardar}")
            infoGuardar.grid(row=5, column=3, padx=30,pady=5)
            infoGuardar.configure(foreground="green", background='#C7E5FC')

            hoja = wb.active  # Puedes cambiar esto para seleccionar una hoja específica

            # Configura el lienzo PDF
            c = canvas.Canvas(f"./_internal/Informes/{nombreArchivoGuardar}.pdf", pagesize=letter)

            # Configura el tamaño de la fuente y el espacio entre líneas
            #c.setFont("Arial", 12)
            espacio_entre_lineas = 14

            # Inicializa las coordenadas para dibujar en la página
            x = 100
            y = 750

            # Itera a través de las filas y columnas del archivo XLSX
            for fila in hoja.iter_rows():
                x = 100
                for celda in fila:
                    # Obtiene el valor de la celda
                    valor = celda.value
                        # # Dibuja el valor en el PDF
                    c.drawString(x, y, str(valor))  # Dibuja en la posición actual (x, y)
                    x+=200
                # Mueve la posición vertical hacia arriba
                y -= espacio_entre_lineas

            # Guarda el PDF
            c.save()

        exportarButton = ttk.Button(informeFrame, text="Guardar en anchivo", command=exportarAExcel)
        exportarButton.grid(row=0, column=3, padx=30, pady=5)
        exportarPdf = ttk.Button(informeFrame, text="Guardar en PDF", command=exportarPdf)
        exportarPdf.grid(row=1, column=3, padx=30, pady=5)


    def muestraAlumnos():
        alumnosFrame = ttk.Frame(frameInforme)
        alumnosFrame.grid()    
        nombresAlumnos = situacion.sheetnames

        InformeGeneralAlumno = tk.Label(alumnosFrame, text="Informe general del alumno")
        InformeGeneralAlumno.grid(column=0, sticky="w", padx=5, pady=5)
        InformeGeneralAlumno.configure(background="#649BEF")

        i = 1
        for nombre in nombresAlumnos[1:]:
            button = [nombre]
            botonAlumno = ttk.Button(alumnosFrame, text=f"Informe de {nombre}",
                                     command= lambda button=button : informeAlumno(button))
            botonAlumno.grid(row=i, column=0, sticky="w", padx=5,pady=1)
            i +=1

        informeGrupoLabel = tk.Label(alumnosFrame, text="Informe general del grupo")
        informeGrupoLabel.grid(row=i+1, column=0, padx=10, pady=20)
        informeGrupoLabel.configure(background="#649BEF")
        button = [nombre]
        informeGrupoButton = ttk.Button(alumnosFrame, text="Informe de todos los alumnos",
                                        command= lambda button=button : informesAllAlumnos())
        informeGrupoButton.grid(row=i+2,column=0)

        InformeRecuperacionAlumno = tk.Label(alumnosFrame, text="Informe de recuperación del alumno")
        InformeRecuperacionAlumno.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        InformeRecuperacionAlumno.configure(background="#649BEF")
        fila = 1
        for nombreRec in nombresAlumnos[1:]:
            button = [nombreRec]
            botonAlumnoRec = ttk.Button(alumnosFrame, text=f"Informe de {nombreRec}",
                                        command=lambda button=button : recuperacionAlumno(button))
            botonAlumnoRec.grid(row=fila, column=1, sticky="w", padx=5,pady=1)
            fila += 1

        informeGrupoRecLabel = tk.Label(alumnosFrame, text="Informe general de recuperación del grupo")
        informeGrupoRecLabel.grid(row=fila+1, column=1, padx=10, pady=20)
        informeGrupoRecLabel.configure(background="#649BEF")
        informeGrupoRecButton = ttk.Button(alumnosFrame, text="Informe de recuperación del grupo",
                                           command=informesDeRecuperacion)
        informeGrupoRecButton.grid(row=fila+2, column=1, padx=10, pady=20)

    muestraAlumnos()

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Informes")
    root.geometry("1000x900")
    root.configure(bg="#649BEF")

    #---------- FRAME PARA METER EL CONTENIDO -------------------------------
    scrollable_frame = moduloScroll.ScrollableFrame(root)
    scrollable_frame.pack(fill="both", expand=True)

    main(root, scrollable_frame.scrollable_frame)

    root.mainloop()
   