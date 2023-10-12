import tkinter as tk, os
from tkinter import ttk
from tkinter import *
import openpyxl, moduloScroll
from tkinter import filedialog
from openpyxl.styles import Font


def cargarSituacion(frameSituacion, infoSituacion):
    infoSituacion.destroy()
    rutaSituacion = filedialog.askopenfilename(title="Eligen la siutación a editar",
                                                initialdir="./_internal/SituacionesAprendizaje/",
                                           filetypes=[("Archivos xlsx", "*.xlsx")])
    infoSituacionB = ttk.Label(frameSituacion, text="Se ha cargado la situación de aprendizaje")
    infoSituacionB.grid(row=6, column=1, padx=10, pady=5)
    infoSituacionB.configure(foreground="green")
    cambiaGrupo(frameSituacion, rutaSituacion)
    

def cambiaGrupo(frameSituacion, rutaSituacion):
    font_negro = Font(color="000000")
    rutaGrupo = filedialog.askopenfilename(title="Eligen un grupo", initialdir="./Grupos/",
                                           filetypes=[("Archivos xlsx", "*.xlsx")])
    if not rutaGrupo:
        infoGrupo = ttk.Label(frameSituacion, text="No se ha cargado ningún grupo, vuelve a elegir grupo")
        infoGrupo.grid(row=3, column=0, padx=10, pady=5)
        infoGrupo.configure(foreground="white")

    grupo = openpyxl.load_workbook(rutaGrupo)
    situacion = openpyxl.load_workbook(rutaSituacion)
    hojaGrupo = grupo.active
    nombresHojas = situacion.sheetnames

    nombres = []
    for fila in hojaGrupo.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        for celda in fila:
            nombres.append(celda)

    print("nombres del grupo", nombres)
    print("nombre de las hojas", nombresHojas)

    # Busca alumnos nuevos 
    alumno = ""
    posicion = 1
    alumnob = nombresHojas[1]
    copiarHoja = situacion[alumnob]
    numeroAlumnos = len(nombres)
    
    for alumno in nombres:
        if alumno not in nombresHojas:
            #copiarHoja.title = alumno
            print(alumno)
            nuevaHoja = situacion.copy_worksheet(copiarHoja)
            # Recorre todas las celdas para cambiar el color
            for fila in nuevaHoja.iter_rows():
                for celda in fila:
                    # Aplica el estilo de fuente negro a la celda
                    celda.font = font_negro

            nuevaHoja.title = alumno
            posicionHoja = (numeroAlumnos-posicion)-2
            situacion.move_sheet(nuevaHoja, -posicionHoja)
            numeroAlumnos +=1
        
            print(numeroAlumnos, posicionHoja, posicion, alumno)
        posicion +=1

    # Elimina hoja de alumno que no está
    for alumnoNo in nombresHojas[1:]:
        print("Alumno:", alumnoNo)
        if alumnoNo not in nombres:
            print("Alumno a eliminar: ", alumnoNo)
            eliminarHoja = situacion[alumnoNo]
            situacion.remove(eliminarHoja)

    print("ruta situacion", rutaSituacion)
    if rutaGrupo and rutaSituacion: 
        cargaLabel = tk.Label(frameSituacion, text=f"Se ha guardado correctamente")
        cargaLabel.grid(row=6, column=0)
        cargaLabel.configure(foreground="green", background="#649BEF")
    else:
        cargaLabel = tk.Label(frameSituacion, text="Falta un dato, no se ha guardado la siutación de aprendizaje")
        cargaLabel.grid(row=5, column=0, sticky="w")
        cargaLabel.configure(foreground="red", background="#649BEF")

    if rutaGrupo:
        infoGrupo = ttk.Label(frameSituacion, text=f"El grupo {rutaGrupo} se ha modificado en la situación de aprendizaje")
        infoGrupo.grid(row=3, column=0, padx=10, pady=5)
        infoGrupo.configure(foreground="white")

    situacion.save(rutaSituacion)
    
def main(root, scrollable_frame):
    rutaSituacion = filedialog.askopenfilename(title="Eligen la siutación a editar",
                                                initialdir="./_internal/SituacionesAprendizaje/",
                                           filetypes=[("Archivos xlsx", "*.xlsx")])
    frameSituacionEdit = ttk.Frame(scrollable_frame)
    frameSituacionEdit.pack()
    infoFrame = ttk.Frame(frameSituacionEdit)
    infoFrame.pack()
    
    frameSituacion = ttk.Frame(frameSituacionEdit)
    frameSituacion.pack(padx=10, pady=5)

    if not rutaSituacion:
        infoSituacion = ttk.Label(frameSituacion, text="No has elegido ninguna situación de aprendizaje, Elige una situacion de aprendizaje")
        infoSituacion.grid(row=6, column=1, padx=10, pady=50)
        infoSituacion.configure(foreground="red")

    style = ttk.Style()
    style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
              background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
    ttk.Style().configure("TFrame", background="#649BEF")
    ttk.Style().configure("TLabel", background="#649BEF", font=(18))
    ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                           borderwidth=0, font=(18), weigth="blod", bordercolor="blue", weight='bold')

    info = tk.Label(infoFrame, text="Selecciona el grupo para modificarlo en la situación de aprendizaje")
    info.grid(row=0, column=0, padx=10, pady=5)
    info.configure(background="#649BEF")
    nombreArchivo = os.path.basename(rutaSituacion)
    infoNombre = tk.Label(frameSituacion, text=f"Situación de aprendizaje: {nombreArchivo}")
    infoNombre.grid(row=2, column=0, sticky='e')
    infoNombre.configure(background="#649BEF")
    grupoButton = ttk.Button(frameSituacion, text="Elige el nuevo Grupo",
                              command=lambda : cambiaGrupo(frameSituacion, rutaSituacion))
    grupoButton.grid(row=4,column=0, padx=10, pady=5)
    situacionButton = ttk.Button(frameSituacion, text="Elige una situacion de aprendizaje",
                              command=lambda : cargarSituacion(frameSituacion, infoSituacion))
    situacionButton.grid(row=4,column=1, padx=10, pady=5)
    if rutaSituacion:
        situacionButton.config(state='disabled', text="Se ha cargado la situación de aprendizaje")

 
if __name__ == "__main__":
    root = tk.Tk()
    root.title("Grupo de alumnos")
    root.geometry("700x1000")
    
    #---------- FRAME PARA METER EL CONTENIDO -------------------------------
    scrollable_frame = moduloScroll.ScrollableFrame(root)
    scrollable_frame.pack(fill="both", expand=True)

    main(root, scrollable_frame.scrollable_frame)

    root.mainloop()
    