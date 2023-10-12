import tkinter as tk
from tkinter import ttk
from tkinter import *
import openpyxl, moduloScroll, os
from tkinter import filedialog

class Situacion():
    def __init__(self, frame, nombreSituacion):
        self.frame = frame
        self.nombre = nombreSituacion

    def crearSituacion(self, rutaGrupo, rutaRubrica, nombreNuevo ):
        # Abre los archivos de entrada
        archivo_nombres = openpyxl.load_workbook(rutaGrupo)
        archivo_datos = openpyxl.load_workbook(rutaRubrica)
        # Crea el archivo de salida
        archivo_salida = openpyxl.Workbook()
        # Selecciona la hoja que desees procesar (por ejemplo, la primera hoja)
        hoja = archivo_nombres.active
        hoja_datos = archivo_datos.active
        # Recorre la primera columna e imprime el contenido de cada celda
        for fila in hoja.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
            for celda in fila:
                # Crea una nueva hoja en el archivo de salida
                nueva_hoja = archivo_salida.create_sheet(title=celda)
            # Copia el contenido de la hoja de datos al nuevo archivo
            for fila_datos in hoja_datos.iter_rows(values_only=True):
                nueva_hoja.append(fila_datos)
        # Guarda el archivo de salida
        archivo_salida.save(f'./_internal/SituacionesAprendizaje/{nombreNuevo}.xlsx')
        # Cierra los archivos
        archivo_nombres.close()
        archivo_datos.close()
   
    def construyeSituacion(self):
        nombreNuevo = self.nombre.get()
        rutaGrupo = filedialog.askopenfilename(title="Eligen un grupo", initialdir="./_internal/Grupos/",
                                               filetypes=[("Archivos xlsx", "*.xlsx")])
        if rutaGrupo:
            nombreGrupo = os.path.basename(rutaGrupo)
            infoGrupo = ttk.Label(self.frame, text=f"El grupo {nombreGrupo} se ha añadido a la situación de aprendizaje")
            infoGrupo.grid(row=3, column=0, padx=10, pady=5)
            infoGrupo.configure(foreground="white")

        rutaRubrica = filedialog.askopenfilename(title="Elige una rúbrica", initialdir="./Rubricas/",
                                                  filetypes=[("Archivos xlsx", "*.xlsx")])
        if rutaRubrica:
            nombreRubrica = os.path.basename(rutaRubrica)
            infoRubrica = ttk.Label(self.frame, text=f"La rúbrica {nombreRubrica} se ha añadido a la situación de aprendizaje")
            infoRubrica.grid(row=4, column=0, padx=10, pady=5)
            infoRubrica.configure(foreground="white")

        if rutaGrupo and rutaRubrica and self.nombre: 
            cargaLabel = tk.Label(self.frame, text=f"Se ha guardado correctamente")
            cargaLabel.grid(row=5, column=0, sticky="w")
            cargaLabel.configure(foreground="green", background="#649BEF")
        else:
            cargaLabel = tk.Label(self.frame, text="Falta un dato, no se ha guardado la siutación de aprendizaje")
            cargaLabel.grid(row=5, column=0, sticky="w")
            cargaLabel.configure(foreground="red", background="#649BEF")

        self.crearSituacion(rutaGrupo, rutaRubrica, nombreNuevo)
        
def main(root, scrollable_frame):
    frameSituacionMain = ttk.Frame(scrollable_frame)
    frameSituacionMain.pack()
    frameSituacion = ttk.Frame(frameSituacionMain)
    frameSituacion.grid(padx=10, pady=5, row=0, column=0, columnspan=1)

    style = ttk.Style()
    style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
              background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
    ttk.Style().configure("TFrame", background="#649BEF")
    ttk.Style().configure("TLabel", background="#649BEF", font=(18))
    ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                           borderwidth=0, font=(18), weigth="blod", bordercolor="blue", weight='bold')
    texto = "Escribe el nombre de la sitiuación de aprendizaje y se abrirán" 
    texto2 = "ventanas para que elijas el curso y la rúbrica."
    texto3 = "- CONSEJO:" 
    texto4 = "Es muy útil que en el nombre de la situación de aprendizaje"
    texto5 = "aparezca el curso, area y situación de aprendizaje"

    etiqueta1 = ttk.Label(frameSituacion, text=texto)
    etiqueta1.grid(row=0, column=0, padx=20, pady=0, sticky='w')
    etiqueta1.configure(background="#649BEF", foreground='black')

    etiqueta2 = ttk.Label(frameSituacion, text=texto2)
    etiqueta2.grid(row=2, column=0, padx=10, pady=0, sticky='w')
    etiqueta2.configure(background="#649BEF", foreground='black')

    etiqueta3 = ttk.Label(frameSituacion, text=texto3)
    etiqueta3.grid(row=3, column=0, padx=20, pady=20, sticky='w')
    etiqueta3.configure(background="#649BEF", foreground='black')

    etiqueta4 = ttk.Label(frameSituacion, text=texto4)
    etiqueta4.grid(row=4, column=0, padx=20, pady=0, sticky='w')
    etiqueta4.configure(background="#649BEF", foreground='black')

    etiqueta5 = ttk.Label(frameSituacion, text=texto5)
    etiqueta5.grid(row=5, column=0, padx=10, pady=0, sticky='w')
    etiqueta5.configure(background="#649BEF", foreground='black')



    infoNombre = ttk.Label(frameSituacion, text="Nombre de la situación de aprendizaje")
    infoNombre.grid(row=0, column=1, padx=200, pady=5)
    nombreSituacion = tk.Entry(frameSituacion)
    nombreSituacion.grid(row=2, column=1, padx=10, pady=5)
    nombreSituacion.configure(width=50)
    situacionNueva = Situacion(frameSituacion, nombreSituacion)
    generaSituacion = ttk.Button(frameSituacion, text="Guardar situación de aprendizaje",
                                 command=lambda: situacionNueva.construyeSituacion())
    generaSituacion.grid(row=3,column=1, padx=10, pady=5)

if __name__ == "__main__":
   root = tk.Tk()
   root.title("Grupo de alumnos")
   root.geometry("700x1000")
   
   #---------- FRAME PARA METER EL CONTENIDO -------------------------------
   scrollable_frame = moduloScroll.ScrollableFrame(root)
   scrollable_frame.pack(fill="both", expand=True)
   main(root, scrollable_frame.scrollable_frame)
   root.mainloop()

    
