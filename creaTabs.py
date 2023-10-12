import tkinter as tk
from tkinter import ttk
from tkinter import *
import openpyxl
from openpyxl.styles import Font


def create_tabs(notebook, situacionAprendizaje, nombresAlumnos, rutaSituacion): 
    tabs = []
    for nombre in nombresAlumnos[1:]:
        tab = ttk.Frame(notebook)
        tabs.append(tab)

        style = ttk.Style()
        style.map("TButton", foreground=[('pressed', 'blue'), ('active', 'black')],
                  background=[('pressed', '!disabled', '#BFD4F3'), ('active', '#C7E5FC')])
        ttk.Style().configure("TFrame", background="#649BEF")
        ttk.Style().configure("TLabel", background="#649BEF", font=(18))
        ttk.Style().configure("TButton", background="#649BEF", foreground="white",
                               borderwidth=0, font=(18), weigth="blod", bordercolor="blue", weight='bold')

        
        notebook.add(tab, text=f"{nombre}")
        ttk.Label(tab, text=f"Alumno: {nombre}").grid(row=0)

        def botonPulsado(boton):
            # Seleccionar el notebook y en sus botones cambiar el color del pulsado
            indice = (nombresAlumnos.index(boton[2]))-1
            
            # Selecciona el tab donde se va a cambiar algo
            notebook.select(tabs[indice])
            
            #Cambia toda la fila de botones y da color al pulsado
            numeroColumnas = len(boton[4])
            hoja = boton[3]
            fila = boton[0]
            button = [boton[0], boton[1], boton[2], boton[3], boton[4]]
            colores =  ['#9C00D5', '#005EFF', '#00C140', '#6FFF00', '#F7FF00', '#FF0000']
            numeroColores = len(colores)

            for columna in range(1,numeroColumnas):
                celda = hoja.cell(row=fila, column=columna+1)
                cell_value = celda.value
                button = [fila, columna+1, boton[2], hoja, boton[4]]
                button = tk.Button(tabs[indice], text=cell_value, width=30, height=10, wraplength=240, 
                                       command=lambda button=button : botonPulsado(button))
                button.configure(borderwidth=0, background='#C7E5FC')
                button.grid(row=fila, column=columna, padx=5, pady=5)
            
                if columna == boton[1]-1:
                    if columna == numeroColumnas-1:
                        button.configure(background=colores[numeroColores-1])
                    elif columna == numeroColumnas-2:
                        button.configure(background=colores[numeroColores-2])
                    elif columna == numeroColumnas-3:
                        button.configure(background=colores[numeroColores-3])
                    elif columna == numeroColumnas-4:
                        button.configure(background=colores[numeroColores-4])
                    elif columna == numeroColumnas-5:
                        button.configure(background=colores[numeroColores-5])
                    else:
                        button.configure(background=colores[numeroColores-6])


            # Cambia el color del elemento pulsado en el archivo xlsx
            letraColumna = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

            # Ruta al archivo xlsx
            archivo_excel = rutaSituacion

            # Cargar el archivo Excel
            libro = openpyxl.load_workbook(archivo_excel)

            # Seleccionar la hoja en la que deseas cambiar la celda
            hoja = libro[boton[2]]
            
            # Selecciona la celda que va a cambiar de color
            fila = boton[0]  # Fila en la que se encuentra la celda
            
            # Modifica el color de toda la fila en la que hemos pulsado botón
            for i in range(1, numeroColumnas+1):
                columna = letraColumna[i-1]  # Columna en la que se encuentra la celda
                cell = hoja[columna + str(fila)]
                font = Font(color="000000")  # Color en formato RGB (verde)
                cell.font = font
                if i == boton[1]:
                    font = Font(color="08CB43")  # Color en formato RGB (verde)
                    cell.font = font

            # Guardar los cambios en el archivo
            libro.save(archivo_excel)
            
            # Cerrar el archivo
            libro.close()
            
        def create_buttons_and_labels(tab, situacionAprendizaje, nombresAlumnos):
            hoja = situacionAprendizaje.active
            row_index = 0
            nombres_columnas = []
            hoja = situacionAprendizaje[nombre]
            numeroFilas = (hoja.max_row)+1
            for columna in hoja.iter_cols(min_row=1, max_row=1, values_only=True):
                nombres_columnas.append(columna)
            # Crear etiquetas y botones en el frame según los datos del archivo
            numeroColumnas = len(nombres_columnas)+1
            colores =  ['#9C00D5', '#005EFF', '#00C140', '#6FFF00', '#F7FF00', '#FF0000']
            numeroColores = len(colores)
            for columna in range(1,numeroColumnas):
                celda = hoja.cell(row=1, column=columna)
                label_text = celda.value
                label = ttk.Label(tab, text=label_text, wraplength=240)
                label.grid(row=row_index+1, column=columna-1,padx=10, pady=5, sticky="w")
                if columna == numeroColumnas-1:
                    label.configure(foreground=colores[numeroColores-1])
                elif columna == numeroColumnas-2:
                    label.configure(foreground=colores[numeroColores-2])
                elif columna == numeroColumnas-3:
                    label.configure(foreground=colores[numeroColores-3])
                elif columna == numeroColumnas-4:
                    label.configure(foreground=colores[numeroColores-4])
                elif columna == numeroColumnas-5:
                    label.configure(foreground=colores[numeroColores-5])
                else:
                    label.configure(foreground=colores[numeroColores-6])
                
            for fila in range(2,numeroFilas):
                for columna in range(1,numeroColumnas):
                    hoja = situacionAprendizaje[nombre]
                    celda = hoja.cell(row=fila, column=columna)
                    cell_value = celda.value
                    if columna ==1:
                        criterio = tk.Label(tab, text=cell_value)
                        criterio.grid(row=fila, column=columna-1, padx=5, pady=5)
                        criterio.configure(background="#649BEF")
                    else:
                        button = [fila, columna, nombre, hoja, nombres_columnas]
                        button = tk.Button(tab, text=cell_value, width=30, height=10, wraplength=240, 
                                           command=lambda button=button : botonPulsado(button))
                        button.configure(borderwidth=0, background='#C7E5FC')
                        button.grid(row=fila, column=columna-1, padx=5, pady=5)
                        fuenteCelda = celda.font
                        color_texto = fuenteCelda.color
                        celdaVerde = 'FF08CB43'
                        if color_texto is not None and color_texto.rgb == celdaVerde or color_texto.rgb == "0008CB43":
                            #button.configure(background="#08CB43")
                            if columna == numeroColumnas-1:
                                button.configure(background=colores[numeroColores-1])
                            elif columna == numeroColumnas-2:
                                button.configure(background=colores[numeroColores-2])
                            elif columna == numeroColumnas-3:
                                button.configure(background=colores[numeroColores-3])
                            elif columna == numeroColumnas-4:
                                button.configure(background=colores[numeroColores-4])
                            elif columna == numeroColumnas-5:
                                button.configure(background=colores[numeroColores-5])
                            else:
                                button.configure(background=colores[numeroColores-6])

    
                
        # Crear etiquetas y botones a partir de los datos del archivo
        create_buttons_and_labels(tab, situacionAprendizaje, nombresAlumnos)